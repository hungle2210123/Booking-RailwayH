"""
Hotel Booking System - Data Backup & Export Module
Provides Excel, CSV, and JSON export functionality for data backup and manual viewing
"""

from flask import send_file, jsonify, request
from datetime import datetime
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from core.models import Booking, Guest, Room, Apartment, db
import json


def register_backup_routes(app):
    """Register all backup and export routes"""

    @app.route('/api/export/excel', methods=['GET'])
    def export_excel():
        """
        Export all booking data to formatted Excel file
        Includes multiple sheets: Bookings, Summary, Revenue
        """
        try:
            # Get date range from query params (optional)
            start_date = request.args.get('start_date')
            end_date = request.args.get('end_date')

            # Query bookings
            query = db.session.query(Booking, Guest, Room, Apartment).outerjoin(
                Guest, Booking.guest_id == Guest.guest_id
            ).outerjoin(
                Room, Booking.room_id == Room.room_id
            ).outerjoin(
                Apartment, Room.apartment_id == Apartment.apartment_id
            ).filter(
                Booking.booking_status != 'deleted'
            )

            # Apply date filters if provided
            if start_date:
                query = query.filter(Booking.checkin_date >= start_date)
            if end_date:
                query = query.filter(Booking.checkin_date <= end_date)

            bookings = query.all()

            # Create Excel workbook
            output = io.BytesIO()
            wb = Workbook()

            # ===== SHEET 1: DETAILED BOOKINGS =====
            ws_bookings = wb.active
            ws_bookings.title = "Chi Tiết Bookings"

            # Headers
            headers = [
                'Mã Booking', 'Tên Khách', 'Email', 'SĐT', 'Quốc Tịch',
                'Check-in', 'Check-out', 'Đêm', 'Phòng/Căn hộ',
                'Tiền Phòng', 'Tiền Taxi', 'Hoa Hồng', 'Đã Thu',
                'Người Thu', 'Trạng Thái', 'Ghi Chú'
            ]

            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for col_num, header in enumerate(headers, 1):
                cell = ws_bookings.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Add booking data
            for row_num, (booking, guest, room, apartment) in enumerate(bookings, 2):
                guest_name = guest.full_name if guest else booking.guest_name or 'N/A'
                nights = (booking.checkout_date - booking.checkin_date).days if booking.checkout_date and booking.checkin_date else 0

                row_data = [
                    booking.booking_id,
                    guest_name,
                    guest.email if guest else '',
                    guest.phone if guest else '',
                    guest.nationality if guest else '',
                    booking.checkin_date.strftime('%Y-%m-%d') if booking.checkin_date else '',
                    booking.checkout_date.strftime('%Y-%m-%d') if booking.checkout_date else '',
                    nights,
                    booking.accommodation_name or room.room_name if room else 'N/A',
                    float(booking.room_amount or 0),
                    float(booking.taxi_amount or 0),
                    float(booking.commission or 0),
                    float(booking.collected_amount or 0),
                    booking.collector or '',
                    booking.booking_status or '',
                    booking.booking_notes or ''
                ]

                for col_num, value in enumerate(row_data, 1):
                    cell = ws_bookings.cell(row=row_num, column=col_num, value=value)

                    # Format currency columns
                    if col_num in [10, 11, 12, 13]:  # Money columns
                        cell.number_format = '#,##0'

                    # Color-code by status
                    if col_num == 15:  # Status column
                        if value in ['OK', 'confirmed', 'checked_in']:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value in ['cancelled', 'đã hủy']:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Auto-adjust column widths
            for col in ws_bookings.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_bookings.column_dimensions[column].width = adjusted_width

            # ===== SHEET 2: SUMMARY STATISTICS =====
            ws_summary = wb.create_sheet("Thống Kê Tổng Quan")

            # Calculate statistics
            total_bookings = len(bookings)
            total_revenue = sum(float(b[0].room_amount or 0) for b in bookings)
            total_collected = sum(float(b[0].collected_amount or 0) for b in bookings)
            total_commission = sum(float(b[0].commission or 0) for b in bookings)
            total_uncollected = total_revenue - total_collected

            stats = [
                ['📊 THỐNG KÊ TỔNG QUAN', ''],
                ['', ''],
                ['Tổng số booking:', total_bookings],
                ['Tổng doanh thu:', total_revenue],
                ['Đã thu:', total_collected],
                ['Chưa thu:', total_uncollected],
                ['Tổng hoa hồng:', total_commission],
                ['', ''],
                ['📅 Ngày xuất báo cáo:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ]

            for row_num, (label, value) in enumerate(stats, 1):
                ws_summary.cell(row=row_num, column=1, value=label).font = Font(bold=True)
                cell = ws_summary.cell(row=row_num, column=2, value=value)
                if isinstance(value, (int, float)) and value > 0:
                    cell.number_format = '#,##0'

            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 20

            # Save to BytesIO
            wb.save(output)
            output.seek(0)

            # Generate filename with timestamp
            filename = f'hotel_bookings_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

        except Exception as e:
            print(f"❌ Excel export error: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500


    @app.route('/api/export/csv', methods=['GET'])
    def export_csv():
        """
        Export booking data to CSV format
        Simple format for universal compatibility
        """
        try:
            # Query bookings
            query = db.session.query(Booking, Guest, Room).outerjoin(
                Guest, Booking.guest_id == Guest.guest_id
            ).outerjoin(
                Room, Booking.room_id == Room.room_id
            ).filter(
                Booking.booking_status != 'deleted'
            )

            bookings = query.all()

            # Create DataFrame
            data = []
            for booking, guest, room in bookings:
                guest_name = guest.full_name if guest else booking.guest_name or 'N/A'
                nights = (booking.checkout_date - booking.checkin_date).days if booking.checkout_date and booking.checkin_date else 0

                data.append({
                    'booking_id': booking.booking_id,
                    'guest_name': guest_name,
                    'email': guest.email if guest else '',
                    'phone': guest.phone if guest else '',
                    'checkin_date': booking.checkin_date.strftime('%Y-%m-%d') if booking.checkin_date else '',
                    'checkout_date': booking.checkout_date.strftime('%Y-%m-%d') if booking.checkout_date else '',
                    'nights': nights,
                    'accommodation': booking.accommodation_name or (room.room_name if room else 'N/A'),
                    'room_amount': float(booking.room_amount or 0),
                    'taxi_amount': float(booking.taxi_amount or 0),
                    'commission': float(booking.commission or 0),
                    'collected_amount': float(booking.collected_amount or 0),
                    'collector': booking.collector or '',
                    'status': booking.booking_status or '',
                    'notes': booking.booking_notes or ''
                })

            df = pd.DataFrame(data)

            # Generate CSV
            output = io.StringIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')  # utf-8-sig for Excel compatibility
            output.seek(0)

            filename = f'hotel_bookings_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

            return send_file(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),
                mimetype='text/csv',
                as_attachment=True,
                download_name=filename
            )

        except Exception as e:
            print(f"❌ CSV export error: {e}")
            return jsonify({'error': str(e)}), 500


    @app.route('/api/export/json', methods=['GET'])
    def export_json():
        """
        Export complete database dump in JSON format
        Useful for system backup and data migration
        """
        try:
            # Query all bookings with full details
            query = db.session.query(Booking, Guest, Room, Apartment).outerjoin(
                Guest, Booking.guest_id == Guest.guest_id
            ).outerjoin(
                Room, Booking.room_id == Room.room_id
            ).outerjoin(
                Apartment, Room.apartment_id == Apartment.apartment_id
            )

            bookings = query.all()

            # Build complete data structure
            export_data = {
                'export_date': datetime.now().isoformat(),
                'total_bookings': len(bookings),
                'bookings': []
            }

            for booking, guest, room, apartment in bookings:
                booking_data = {
                    'booking_id': booking.booking_id,
                    'guest_name': guest.full_name if guest else booking.guest_name,
                    'guest_email': guest.email if guest else None,
                    'guest_phone': guest.phone if guest else None,
                    'guest_nationality': guest.nationality if guest else None,
                    'checkin_date': booking.checkin_date.isoformat() if booking.checkin_date else None,
                    'checkout_date': booking.checkout_date.isoformat() if booking.checkout_date else None,
                    'room_amount': float(booking.room_amount or 0),
                    'taxi_amount': float(booking.taxi_amount or 0),
                    'commission': float(booking.commission or 0),
                    'collected_amount': float(booking.collected_amount or 0),
                    'collector': booking.collector,
                    'accommodation_name': booking.accommodation_name,
                    'room_name': room.room_name if room else None,
                    'apartment_name': apartment.name if apartment else None,
                    'booking_status': booking.booking_status,
                    'booking_notes': booking.booking_notes,
                    'booking_platform': booking.booking_platform,
                    'created_at': booking.created_at.isoformat() if hasattr(booking, 'created_at') and booking.created_at else None
                }
                export_data['bookings'].append(booking_data)

            # Create JSON file
            filename = f'hotel_bookings_full_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'

            output = io.BytesIO()
            output.write(json.dumps(export_data, indent=2, ensure_ascii=False).encode('utf-8'))
            output.seek(0)

            return send_file(
                output,
                mimetype='application/json',
                as_attachment=True,
                download_name=filename
            )

        except Exception as e:
            print(f"❌ JSON export error: {e}")
            return jsonify({'error': str(e)}), 500
