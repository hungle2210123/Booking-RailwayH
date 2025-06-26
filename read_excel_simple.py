#!/usr/bin/env python3
"""
Read Excel file using openpyxl to understand template structure
"""
try:
    from openpyxl import load_workbook
    
    def read_excel_templates():
        try:
            # Read the Excel file
            file_path = "/mnt/c/Users/T14/Desktop/hotel_flask_app/hotel_flask_app_optimized/csvtest (1).xlsx"
            
            wb = load_workbook(file_path, read_only=True)
            print(f"📊 Excel file sheets: {wb.sheetnames}")
            
            # Read each sheet
            for sheet_name in wb.sheetnames:
                print(f"\n📋 Reading sheet: {sheet_name}")
                ws = wb[sheet_name]
                
                # Get headers
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value)
                print(f"Headers: {headers}")
                
                # Get first 10 rows of data
                print("First 10 rows:")
                for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if i <= 10:
                        print(f"Row {i}: {row}")
                    else:
                        break
                
                print(f"Total rows: {ws.max_row}")
                print("\n" + "="*80)
            
        except Exception as e:
            print(f"Error reading Excel file: {e}")
    
    if __name__ == "__main__":
        read_excel_templates()

except ImportError:
    print("openpyxl not available, trying with xlrd...")
    
    try:
        import xlrd
        
        def read_excel_templates():
            try:
                file_path = "/mnt/c/Users/T14/Desktop/hotel_flask_app/hotel_flask_app_optimized/csvtest (1).xlsx"
                
                workbook = xlrd.open_workbook(file_path)
                print(f"📊 Excel file sheets: {workbook.sheet_names()}")
                
                for sheet_name in workbook.sheet_names():
                    print(f"\n📋 Reading sheet: {sheet_name}")
                    sheet = workbook.sheet_by_name(sheet_name)
                    
                    # Get headers
                    if sheet.nrows > 0:
                        headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
                        print(f"Headers: {headers}")
                        
                        # Get first 10 rows
                        print("First 10 rows:")
                        for row in range(min(10, sheet.nrows)):
                            row_data = [sheet.cell_value(row, col) for col in range(sheet.ncols)]
                            print(f"Row {row+1}: {row_data}")
                        
                        print(f"Total rows: {sheet.nrows}")
                    print("\n" + "="*80)
                        
            except Exception as e:
                print(f"Error reading Excel file: {e}")
        
        if __name__ == "__main__":
            read_excel_templates()
            
    except ImportError:
        print("Neither openpyxl nor xlrd available. Cannot read Excel file.")
        print("Please provide the Excel file content in text format or CSV.")